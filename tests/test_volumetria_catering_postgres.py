"""Volumetria de catering — a prova contra Postgres REAL.

O SQL da consulta é dialeto Postgres (`= ANY(array)`, `to_char`, `EXTRACT`,
`AT TIME ZONE`, cursor nomeado); a suíte do Hub roda SQLite. A fricção foi
aceita pela Maria em 27/ago/2026: este arquivo exige um Postgres de teste
DEDICADO e é **pulado inteiro** (skip, não falha) quando ele não responde.

Banco esperado (ver docs/EXECUCAO_LOCAL.md — sobe uma vez, fica parado entre
sessões):

    wsl -d Ubuntu-24.04 -e docker run -d --name superfrio-teste-db --memory=256m \
      -p 5434:5432 -e POSTGRES_USER=hub_teste -e POSTGRES_PASSWORD=teste \
      -e POSTGRES_DB=hub_teste postgres:16-alpine -c shared_buffers=32MB

Outra URL: `VOLUMETRIA_TEST_DB_URL`. **Cada teste zera o schema `public`
(DROP SCHEMA CASCADE)** — nunca aponte para um banco com dado de verdade. Por
isso NÃO se reaproveita o `nuvem-teste-db` da nuvem-ia (porta 5433).

O que este arquivo prova, e o SQLite não conseguiria:

- a Matriz agrega, e o Decimal sai como STRING no JSON;
- a soma de todas as páginas da planilha bate com a Matriz no mesmo recorte —
  o aceite do V3.3, agora com o driver do Hub;
- o CSV é Excel-first (BOM, `;`, vírgula decimal, `DD/MM/AAAA`) e o xlsx
  escreve identificador como TEXTO;
- `pagina` não entra no download; `amb` soma entrada e saída só na Matriz;
- o filtro de dia do mês recorta dentro do mês (é onde o `ANY(int[])` do
  psycopg 3 é exercitado de verdade);
- a verificação de drift falha nomeando a coluna, para coluna que sumiu, coluna
  nova e tipo trocado — e o download não abre auditoria quando falha;
- a conexão do módulo é SOMENTE LEITURA por mecanismo (`ReadOnlySqlTransaction`);
- `ver` libera a consulta e `exportar` libera o download, ponta a ponta;
- a auditoria no banco do Hub guarda o recorte e a contagem real.
"""
import io
import os
import socket
from decimal import Decimal
from pathlib import Path

import psycopg
import pytest

from backend.volumetria_catering import auditoria, conexao, schema
from backend.volumetria_catering.permissoes import APP_SLUG, EXPORTAR

URL = os.environ.get(
    "VOLUMETRIA_TEST_DB_URL", "postgresql://hub_teste:teste@localhost:5434/hub_teste"
)


def _alcancavel() -> bool:
    """Postgres de teste responde? Em duas etapas, com timeout duro nas duas.

    A etapa de socket existe por causa do relay de porta do WSL: quando a distro
    encerra sozinha (~8 s sem processo `wsl` nem conexão — ver
    docs/EXECUCAO_LOCAL.md §3.1) o `localhost:5434` do Windows pode aceitar o
    TCP e nunca responder, e um `connect` do driver ficou preso por mais de 10
    minutos numa rodada (27/08/2026). Um socket cru com timeout não fica."""
    try:
        partes = psycopg.conninfo.conninfo_to_dict(URL)
        host = partes.get("host") or "localhost"
        porta = int(partes.get("port") or 5432)
        with socket.create_connection((host, porta), timeout=2):
            pass
        with psycopg.connect(URL, connect_timeout=2):
            return True
    except (OSError, psycopg.Error, ValueError):
        return False


def url_de_teste_local(url: str) -> bool:
    """A trava antes do `DROP SCHEMA`: só banco LOCAL com 'teste' no nome do
    banco E do usuário. Ação destrutiva exige alvo explícito (política §6.5) —
    um túnel para o `nuvem-db` ou um dump com dado real não pode ser zerado
    por uma variável exportada por distração."""
    try:
        partes = psycopg.conninfo.conninfo_to_dict(url)
    except psycopg.ProgrammingError:
        return False
    host = (partes.get("host") or "localhost").lower()
    banco = (partes.get("dbname") or "").lower()
    usuario = (partes.get("user") or "").lower()
    return host in ("localhost", "127.0.0.1", "::1") and "teste" in banco and "teste" in usuario


if not _alcancavel():
    pytest.skip(
        f"Postgres de teste da volumetria indisponível em {URL.split('@')[-1]} "
        "— ver docs/EXECUCAO_LOCAL.md (seção Postgres de teste da volumetria)",
        allow_module_level=True,
    )

SCHEMA_SQL = (Path(__file__).parent / "volumetria_catering_schema.sql").read_text(encoding="utf-8")
BASE = "/api/volumetria-catering"
JAN = {"de": "2026-01-01", "ate": "2026-01-31"}
TABELA_REC = "DM_VOLUMETRIA.FATO_VOL_REC_CAT_V01"
TABELA_EXP = "DM_VOLUMETRIA.FATO_VOL_EXP_CAT_V01"


@pytest.fixture
def cat(monkeypatch):
    """Schema `cat_*` zerado e recriado do DDL; o módulo apontado para ele.

    Devolve uma conexão ESCREVÍVEL (autocommit) para semear — a do módulo é
    somente leitura, e é isso que um dos testes prova."""
    if not url_de_teste_local(URL):
        pytest.fail(
            f"VOLUMETRIA_TEST_DB_URL recusada ({URL.split('@')[-1]}): esta suíte ZERA o "
            "schema public, então só aceita host local e banco/usuário com 'teste' no nome."
        )
    monkeypatch.setenv(conexao.ENV_URL, URL)
    schema.invalidar()
    conn = psycopg.connect(URL, autocommit=True)
    with conn.cursor() as cur:
        cur.execute("DROP SCHEMA public CASCADE")
        cur.execute("CREATE SCHEMA public")
        cur.execute(SCHEMA_SQL)
    yield conn
    conn.close()
    schema.invalidar()


# ------------------------------------------------------------- semeadura
# Porte dos semeadores de `tests/test_catering_matriz.py` da nuvem-ia.
_COMUNS = (
    "pk_dw, dw_processo, dw_data_inclusao, dw_data_alteracao, sk_calendario,"
    " sk_instancia, sk_empresa, sk_filial, sk_cliente, nk_calendario,"
    " nk_instancia, nk_empresa, nk_filial, nk_wms_filial, nk_qls_filial,"
    " nk_slin_empresa, nk_slin_filial, nk_cliente, nk_wms_cliente, data_solic,"
    " ano_solic, nome_und, num_gem, cnpj_cpf_cli, raz_social, descr_oper_wms,"
    " nome_estoque, status_processo, flg_interface"
)


def _carga(conn, tabela, linhas=1):
    with conn.cursor() as cur:
        cur.execute(
            "INSERT INTO cat_cargas (tabela_origem, fonte, status, terminada_em, linhas_lidas) "
            "VALUES (%s, 'oracle', 'ok', now(), %s) RETURNING id",
            (tabela, linhas),
        )
        return cur.fetchone()[0]


def _valores(sigla, cliente, gem, operacao, calendario, solic):
    return (
        1, "catering_to_dw_volumetry_v01", "2026-08-20 15:00:00", "2026-08-20 15:00:00",
        1, 1, 1, 1, 1, calendario, "SLIN_RMSPII_PRD", "SF", "06975242000187",
        sigla, sigla, "001", "001", cliente, "X", solic, 2026,
        f"{sigla} - TESTE", gem, f"{cliente}0001", "CLIENTE TESTE", operacao,
        "CONGELADO", "Concluido", "D",
    )


def semear_entrada(conn, sigla="RMSPII", cliente="67945071", gem="0000000001",
                   operacao="NAO TROCA NOTA DE ARMAZENAGEM", peso="100.000",
                   calendario="2026-01-05", solic="2026-01-05"):
    carga = _carga(conn, TABELA_REC)
    with conn.cursor() as cur:
        cur.execute(
            f"INSERT INTO cat_fato_recebimento (carga_id, {_COMUNS},"
            " qtde_sku, qtde_pallet, qtde_vol2, qtde_peso2, qtde_pbrt2, qtde_vlr)"
            " VALUES (%s" + ", %s" * 29 + ", 1, 7, 10, %s, %s, %s)",
            (carga,) + _valores(sigla, cliente, gem, operacao, calendario, solic)
            + (peso, peso, peso),
        )


def semear_saida(conn, sigla="RMSPII", cliente="67945071", gem="0000000001",
                 operacao="SAIDA NORMAL", calendario="2026-01-05"):
    """Solicitado 100, atendido 80, separado 70 — três faixas com valores
    distintos, para dar para provar qual delas a tela mostrou."""
    carga = _carga(conn, TABELA_EXP)
    with conn.cursor() as cur:
        cur.execute(
            f"INSERT INTO cat_fato_expedicao (carga_id, {_COMUNS},"
            " qtde_pedido,"
            " qtde_sku_solicitado, qtde_vol_solicitado, qtde_peso_solicitado,"
            " qtde_pbrt_solicitado, qtde_vlr_solicitado,"
            " qtde_sku_atendido, qtde_vol_atendido, qtde_peso_atendido,"
            " qtde_pbrt_atendido, qtde_vlr_atendido,"
            " qtde_sku_separado, qtde_vol_separado, qtde_peso_separado,"
            " qtde_pbrt_separado, qtde_vlr_separado)"
            " VALUES (%s" + ", %s" * 29 + ", 1,"
            " 1, 10, 100.000, 100.000, 100.000,"
            " 1, 8, 80.000, 80.000, 80.000,"
            " 1, 7, 70.000, 70.000, 70.000)",
            (carga,) + _valores(sigla, cliente, gem, operacao, calendario, calendario),
        )


def _linhas_csv(conteudo: bytes) -> list[str]:
    texto = conteudo.decode("utf-8-sig")
    return [l for l in texto.split("\r\n") if l]


# ============ A trava da fixture ============

@pytest.mark.parametrize("url, aceita", [
    ("postgresql://hub_teste:teste@localhost:5434/hub_teste", True),
    ("postgresql://hub_teste:teste@127.0.0.1:5434/hub_teste", True),
    ("postgresql://nuvem:teste@localhost:5433/nuvem_teste", False),   # usuário sem 'teste'
    ("postgresql://hub_teste:x@localhost:5434/nuvem", False),         # banco sem 'teste'
    ("postgresql://hub_teste:x@nuvem-db:5432/hub_teste", False),      # host remoto
    ("postgresql://hub_teste:x@10.0.0.5:5432/hub_teste", False),
    ("isto nao e uma url", False),
])
def test_fixture_so_zera_banco_de_teste_local(url, aceita):
    assert url_de_teste_local(url) is aceita


# ============ Contrato e conexão ============

def test_contrato_copiado_bate_com_o_schema_da_nuvem_ia(cat):
    """O DDL de teste é a forma final das migrations 0019–0024. Se este teste
    quebrar, ou a cópia do contrato ou a cópia do DDL ficou para trás."""
    conn = conexao.conectar()
    try:
        with conn.cursor() as cur:
            schema.verificar(cur)
    finally:
        conn.close()


def test_conexao_do_modulo_e_somente_leitura_por_mecanismo(cat):
    """Mesmo com um usuário dono de tudo na URL, escrever pela conexão do
    módulo é recusado pelo Postgres — `default_transaction_read_only=on`."""
    conn = conexao.conectar()
    try:
        with conn.cursor() as cur:
            with pytest.raises(psycopg.errors.ReadOnlySqlTransaction):
                cur.execute("INSERT INTO cat_cargas (tabela_origem) VALUES ('x')")
    finally:
        conn.close()
    # o usuário de teste PODE escrever por fora: a trava é da conexão do módulo
    _carga(cat, TABELA_REC)


# ============ Opções e procedência ============

def test_opcoes_saem_do_dado_com_procedencia(cat, client, admin_headers):
    semear_entrada(cat)
    semear_entrada(cat, sigla="XPTO", cliente="99999999", gem="0000000002",
                   operacao="OPERACAO NOVA", calendario="2026-02-10", solic="2026-02-10")
    semear_saida(cat)

    r = client.get(f"{BASE}/opcoes", headers=admin_headers)
    assert r.status_code == 200, r.text
    corpo = r.json()
    assert corpo["unidades"] == ["RMSPII", "XPTO"]  # sem cat_unidades: sigla da fonte
    assert {"67945071", "99999999"} == {c["chave"] for c in corpo["clientes"]}
    assert "OPERACAO NOVA" in corpo["operacoes"]["rec"]
    assert corpo["operacoes"]["exp"] == ["SAIDA NORMAL"]
    assert corpo["periodo"] == {"de": "2026-01-05", "ate": "2026-02-10"}
    assert corpo["abertura"]["de"].endswith("-01-01")
    assert corpo["abertura"]["de"] <= corpo["abertura"]["ate"]
    assert corpo["teto_xlsx"] == 150_000
    # procedência: as duas últimas cargas ok, com quando e quantas linhas
    assert len(corpo["cargas"]) == 2
    assert {c["tabela"] for c in corpo["cargas"]} == {TABELA_REC, TABELA_EXP}
    assert all(c["quando"] and c["fonte"] == "oracle" for c in corpo["cargas"])
    assert [m["chave"] for m in corpo["movimentos"]] == ["rec", "exp", "amb"]


# ============ Matriz ============

def test_matriz_agrega_e_devolve_decimal_como_string(cat, client, admin_headers):
    semear_entrada(cat, gem="0000000001", peso="100.000")
    semear_entrada(cat, gem="0000000002", peso="250.500")

    r = client.get(f"{BASE}/matriz", params=JAN, headers=admin_headers)
    assert r.status_code == 200, r.text
    corpo = r.json()
    assert corpo["total"] == {"2026-01": "350.500"}  # string, não float
    assert corpo["total_linhas"] == 2
    assert corpo["niveis"] == ["unidade", "cliente", "operacao"]
    unidade = corpo["linhas"][0]
    assert unidade["chave"] == "RMSPII"
    assert unidade["valores"] == {"2026-01": "350.500"}
    assert unidade["filhos"][0]["chave"] == "67945071"
    assert unidade["filhos"][0]["rotulo"] == "CLIENTE TESTE"


def test_matriz_conjunta_soma_entrada_e_saida_pela_faixa_escolhida(cat, client, admin_headers):
    semear_entrada(cat, peso="100.000")
    semear_saida(cat)

    r = client.get(f"{BASE}/matriz", params={**JAN, "movimento": "amb", "faixa": "atendido"},
                   headers=admin_headers)
    assert r.status_code == 200, r.text
    corpo = r.json()
    assert corpo["niveis"] == ["unidade", "cliente", "movimento"]
    assert corpo["total"] == {"2026-01": "180.000"}  # 100 entrada + 80 atendido
    cliente = corpo["linhas"][0]["filhos"][0]
    assert [f["chave"] for f in cliente["filhos"]] == ["exp", "rec"]  # ordem fixa
    assert any("movimentação" in a for a in corpo["avisos"])


def test_filtro_de_dia_do_mes_recorta_dentro_do_mes(cat, client, admin_headers):
    semear_entrada(cat, gem="0000000001", calendario="2026-01-05", solic="2026-01-05")
    semear_entrada(cat, gem="0000000002", calendario="2026-01-20", solic="2026-01-20")

    r = client.get(f"{BASE}/matriz", params={**JAN, "dia": ["5", "6"]}, headers=admin_headers)
    assert r.status_code == 200, r.text
    corpo = r.json()
    assert corpo["total"] == {"2026-01": "100.000"}
    assert corpo["total_linhas"] == 1
    assert corpo["filtros"]["dias"] == [5, 6]
    assert any("Filtro de dia do mês" in a for a in corpo["avisos"])


def test_dimensoes_canonizam_e_filtros_de_cliente_tipo_e_operacao(cat, client, admin_headers):
    """As três dimensões de decisão da nuvem-ia entram por LEFT JOIN + COALESCE:
    a sigla exibida (RMSPV → RMSPIV), a razão social canônica e o tipo de
    estoque. E os filtros que a tela manda por elas têm que estreitar o recorte
    — inclusive o de unidade, que filtra pela sigla EXIBIDA."""
    with cat.cursor() as cur:
        cur.execute("INSERT INTO cat_unidades VALUES ('RMSPV', 'RMSPIV', 'Unidade IV', now())")
        cur.execute("INSERT INTO cat_clientes (raiz_cnpj, razao_social) VALUES ('67945071', 'CLIENTE CANONICO SA')")
        cur.execute("INSERT INTO cat_tipos_estoque (nome_estoque, tipo) VALUES ('CONGELADO', 'CONGELADO')")
    semear_entrada(cat, sigla="RMSPV", cliente="67945071", gem="0000000001")
    semear_entrada(cat, sigla="RMSPII", cliente="99999999", gem="0000000002", operacao="DEVOLUCAO")

    corpo = client.get(f"{BASE}/matriz", params=JAN, headers=admin_headers).json()
    unidades = {u["chave"]: u for u in corpo["linhas"]}
    assert set(unidades) == {"RMSPIV", "RMSPII"}  # sigla exibida, com queda para a fonte
    assert unidades["RMSPIV"]["filhos"][0]["rotulo"] == "CLIENTE CANONICO SA"
    assert unidades["RMSPII"]["filhos"][0]["rotulo"] == "CLIENTE TESTE"  # sem dimensão: grafia da linha

    def total(**extra):
        return client.get(f"{BASE}/matriz", params={**JAN, **extra}, headers=admin_headers).json()["total"]["2026-01"]

    assert total(unidade=["RMSPIV"]) == "100.000"       # filtra pela sigla EXIBIDA
    assert total(cliente=["99999999"]) == "100.000"
    assert total(operacao=["DEVOLUCAO"]) == "100.000"
    assert total(tipo_estoque=["CONGELADO"]) == "200.000"
    assert total(tipo_estoque=["SECO"]) is None          # nada casa: coluna vazia, não erro

    planilha = client.get(f"{BASE}/planilha", params={**JAN, "cliente": ["67945071"]}, headers=admin_headers).json()
    assert [l["tipo_estoque"] for l in planilha["linhas"]] == ["CONGELADO"]
    assert planilha["linhas"][0]["unidade"] == "RMSPIV"


def test_xlsx_acima_do_teto_recusa_com_400_e_marca_a_auditoria(cat, client, admin_headers, monkeypatch):
    semear_entrada(cat, gem="0000000001")
    semear_entrada(cat, gem="0000000002")
    monkeypatch.setattr("backend.volumetria_catering.download.TETO_XLSX", 1)
    r = client.get(f"{BASE}/download", params={**JAN, "formato": "xlsx"}, headers=admin_headers)
    assert r.status_code == 400
    assert "CSV" in r.json()["detail"]
    ultimo = auditoria.listar(1)[0]
    assert ultimo["status"] == "erro" and ultimo["formato"] == "xlsx"


def test_filtro_de_unidade_e_periodo_parcial(cat, client, admin_headers):
    semear_entrada(cat, sigla="RMSPII", gem="0000000001")
    semear_entrada(cat, sigla="XPTO", gem="0000000002", calendario="2026-01-25", solic="2026-01-25")

    r = client.get(f"{BASE}/matriz", params={"de": "2026-01-03", "ate": "2026-01-31", "unidade": ["XPTO"]},
                   headers=admin_headers)
    corpo = r.json()
    assert [u["chave"] for u in corpo["linhas"]] == ["XPTO"]
    assert corpo["rotulos_meses"] == {"2026-01": "2026-01 (03-31)"}


# ============ Planilha ============

def test_planilha_somada_bate_com_a_matriz(cat, client, admin_headers):
    """O aceite do V3.3: agregação e detalhe contam as MESMAS linhas."""
    for i, peso in enumerate(("100.000", "250.500", "0.250"), start=1):
        semear_entrada(cat, gem=f"{i:010d}", peso=peso)

    matriz = client.get(f"{BASE}/matriz", params=JAN, headers=admin_headers).json()
    planilha = client.get(f"{BASE}/planilha", params=JAN, headers=admin_headers).json()

    assert planilha["paginacao"]["total_linhas"] == matriz["total_linhas"] == 3
    soma = sum(Decimal(l["valor"]) for l in planilha["linhas"])
    assert soma == Decimal(matriz["total"]["2026-01"])
    assert [c["chave"] for c in planilha["colunas"]] == [
        "dia", "unidade", "cliente", "guia", "operacao", "tipo_estoque", "valor",
    ]
    assert planilha["linhas"][0]["guia"].startswith("000000000")  # zero à esquerda intacto


def test_planilha_pagina_alem_do_fim_avisa_e_nao_erra(cat, client, admin_headers):
    semear_entrada(cat)
    r = client.get(f"{BASE}/planilha", params={**JAN, "pagina": 7}, headers=admin_headers)
    assert r.status_code == 200
    assert r.json()["linhas"] == []
    assert any("além do fim" in a for a in r.json()["avisos"])


def test_planilha_da_saida_traz_as_tres_faixas(cat, client, admin_headers):
    semear_saida(cat)
    r = client.get(f"{BASE}/planilha", params={**JAN, "movimento": "exp"}, headers=admin_headers)
    linha = r.json()["linhas"][0]
    assert (linha["solicitado"], linha["atendido"], linha["separado"]) == ("100.000", "80.000", "70.000")


# ============ Download ============

def test_download_csv_excel_first_e_com_o_recorte_inteiro(cat, client, admin_headers):
    semear_entrada(cat, gem="0000000001", peso="100.000")
    semear_entrada(cat, gem="0000000002", peso="250.500")
    semear_entrada(cat, gem="0000000003", peso="1.000", calendario="2026-02-01", solic="2026-02-01")

    # `pagina` é ignorado: download nunca é de uma página só
    r = client.get(f"{BASE}/download", params={**JAN, "pagina": 2}, headers=admin_headers)
    assert r.status_code == 200, r.text
    assert r.headers["content-type"].startswith("text/csv")
    assert 'filename="catering_entrada_2026-01-01_a_2026-01-31.csv"' in r.headers["content-disposition"]
    assert r.content.startswith(b"\xef\xbb\xbf")  # BOM: o Excel abre os acentos certos

    linhas = _linhas_csv(r.content)
    cabecalho = linhas[0].split(";")
    assert cabecalho[:4] == ["Dia", "Unidade", "Cliente", "Tipo de estoque"]
    assert len(cabecalho) == 4 + 36
    assert len(linhas) == 1 + 2  # só janeiro
    campos = dict(zip(cabecalho, linhas[1].split(";")))
    assert campos["Dia"] == "05/01/2026"
    assert campos["num_gem"] == "0000000001"  # zero à esquerda intacto no CSV
    assert campos["qtde_peso2"] == "100,000"  # vírgula decimal
    assert campos["Unidade"] == "RMSPII"
    assert campos["nk_filial"] == "06975242000187"


def test_download_xlsx_escreve_identificador_como_texto(cat, client, admin_headers):
    from openpyxl import load_workbook

    semear_entrada(cat, gem="0000000609")
    r = client.get(f"{BASE}/download", params={**JAN, "formato": "xlsx"}, headers=admin_headers)
    assert r.status_code == 200, r.text
    assert "spreadsheetml" in r.headers["content-type"]

    aba = load_workbook(io.BytesIO(r.content), read_only=True)["volumetria"]
    linhas = list(aba.iter_rows(values_only=False))
    cabecalho = [c.value for c in linhas[0]]
    idx_gem = cabecalho.index("num_gem")
    idx_peso = cabecalho.index("qtde_peso2")
    celula_gem = linhas[1][idx_gem]
    assert celula_gem.value == "0000000609"
    assert celula_gem.number_format == "@"
    assert isinstance(linhas[1][idx_peso].value, (int, float, Decimal))  # medida continua número


def test_download_exige_exportar_e_ver_libera_so_a_consulta(cat, client, admin_headers):
    semear_entrada(cat)
    r = client.post("/api/admin/roles", json={"slug": "vol-pg-ver", "nome": "V", "apps": [APP_SLUG]},
                    headers=admin_headers)
    assert r.status_code == 201, r.text
    role_id = r.json()["id"]
    r = client.post("/api/admin/usuarios",
                    json={"username": "vol.pg", "senha": "senha-de-teste-123", "roles": ["vol-pg-ver"]},
                    headers=admin_headers)
    assert r.status_code == 201, r.text
    token = client.post("/api/auth/login", data={"username": "vol.pg", "password": "senha-de-teste-123"}).json()["access_token"]
    headers = {"Authorization": f"Bearer {token}"}

    assert client.get(f"{BASE}/matriz", params=JAN, headers=headers).status_code == 200
    assert client.get(f"{BASE}/download", params=JAN, headers=headers).status_code == 403

    # a célula `exportar` da matriz libera
    r = client.patch(f"/api/admin/roles/{role_id}", json={"permissoes": [EXPORTAR]}, headers=admin_headers)
    assert r.status_code == 200, r.text
    r = client.get(f"{BASE}/download", params=JAN, headers=headers)
    assert r.status_code == 200
    assert len(_linhas_csv(r.content)) == 2


def test_auditoria_registra_quem_baixou_o_que_com_a_contagem_real(cat, client, admin_headers):
    semear_entrada(cat, gem="0000000001")
    semear_entrada(cat, gem="0000000002")
    params = {**JAN, "unidade": ["RMSPII"], "dia": ["5"]}
    r = client.get(f"{BASE}/download", params=params, headers=admin_headers)
    assert r.status_code == 200
    assert len(_linhas_csv(r.content)) == 1 + 2

    ultimo = auditoria.listar(1)[0]
    assert ultimo["status"] == "ok"
    assert ultimo["linhas"] == 2
    assert ultimo["usuario"] == "admin"
    assert ultimo["formato"] == "csv"
    assert ultimo["ip"]
    assert ultimo["recorte"]["unidades"] == ["RMSPII"]
    assert ultimo["recorte"]["dias"] == [5]
    assert ultimo["recorte"]["de"] == "2026-01-01"

    api = client.get(f"{BASE}/auditoria", params={"limite": 1}, headers=admin_headers).json()
    assert api[0]["id"] == ultimo["id"]


# ============ Drift contra o banco de verdade ============

def _drop_coluna(conn, tabela, coluna):
    with conn.cursor() as cur:
        cur.execute(f"ALTER TABLE {tabela} DROP COLUMN {coluna}")


def test_drift_coluna_removida_derruba_o_card_nomeando_a_coluna(cat, client, admin_headers):
    semear_entrada(cat)
    # antes do drift, funciona (e o resultado bom entra no cache)
    assert client.get(f"{BASE}/matriz", params=JAN, headers=admin_headers).status_code == 200

    _drop_coluna(cat, "cat_fato_recebimento", "qtde_pbrt2")
    schema.invalidar()  # em produção, o cache vence em 10 min
    antes = len(auditoria.listar(1000))

    r = client.get(f"{BASE}/matriz", params=JAN, headers=admin_headers)
    assert r.status_code == 503
    assert "cat_fato_recebimento.qtde_pbrt2" in r.json()["detail"]
    assert "contrato.py" in r.json()["detail"]
    # download também para — e ANTES de abrir auditoria
    r = client.get(f"{BASE}/download", params=JAN, headers=admin_headers)
    assert r.status_code == 503
    assert len(auditoria.listar(1000)) == antes
    # o resto do Hub não sente
    assert client.get("/api/health").status_code == 200


def test_drift_coluna_nova_na_nuvem_ia_tambem_e_drift(cat, client, admin_headers):
    with cat.cursor() as cur:
        cur.execute("ALTER TABLE cat_fato_expedicao ADD COLUMN qtde_nova INTEGER")
    r = client.get(f"{BASE}/opcoes", headers=admin_headers)
    assert r.status_code == 503
    assert "cat_fato_expedicao.qtde_nova" in r.json()["detail"]


def test_drift_tipo_trocado(cat, client, admin_headers):
    with cat.cursor() as cur:
        cur.execute("ALTER TABLE cat_fato_recebimento ALTER COLUMN qtde_sku TYPE BIGINT")
    r = client.get(f"{BASE}/planilha", params=JAN, headers=admin_headers)
    assert r.status_code == 503
    assert "cat_fato_recebimento.qtde_sku: tipo esperado INTEGER, banco tem bigint" in r.json()["detail"]


def test_verificacao_boa_fica_em_cache_e_falha_nao(cat, client, admin_headers):
    """Uma verificação boa vale por `INTERVALO_REVERIFICACAO`; uma falha é
    reconferida no request seguinte — corrigir o banco basta, sem reiniciar."""
    assert client.get(f"{BASE}/opcoes", headers=admin_headers).status_code == 200
    with cat.cursor() as cur:
        cur.execute("ALTER TABLE cat_fato_recebimento ADD COLUMN passageira INTEGER")
    # cache ainda válido: não percebe (é o custo aceito dos 10 min)
    assert client.get(f"{BASE}/opcoes", headers=admin_headers).status_code == 200
    schema.invalidar()
    assert client.get(f"{BASE}/opcoes", headers=admin_headers).status_code == 503
    # sem cache de falha: corrigido o banco, o próximo request passa
    _drop_coluna(cat, "cat_fato_recebimento", "passageira")
    assert client.get(f"{BASE}/opcoes", headers=admin_headers).status_code == 200


def test_download_pelo_ticket_sai_igual_e_audita_quem_baixou(cat, client, admin_headers):
    """O caminho REAL da tela, ponta a ponta: pede o ticket com o Bearer e baixa
    por navegação, sem header nenhum.

    É o único caminho que o navegador tem (`ticket.py` explica por quê), então
    ele precisa entregar exatamente o mesmo arquivo do caminho por Bearer — e a
    auditoria precisa gravar o usuário certo, porque é ela que responde quem
    baixou o quê.
    """
    semear_entrada(cat, gem="0000000001", peso="100.000")
    semear_entrada(cat, gem="0000000002", peso="250.500")

    r = client.post(f"{BASE}/download/ticket", params=JAN, headers=admin_headers)
    assert r.status_code == 200, r.text
    corpo = r.json()
    assert corpo["valido_por_segundos"] == 60

    # a navegação NÃO leva Authorization — é o ticket que prova quem é
    baixado = client.get(f"{BASE}/download", params={**JAN, "ticket": corpo["ticket"]})
    assert baixado.status_code == 200, baixado.text
    assert baixado.headers["content-type"].startswith("text/csv")
    assert 'filename="catering_entrada_2026-01-01_a_2026-01-31.csv"' in \
        baixado.headers["content-disposition"]

    por_bearer = client.get(f"{BASE}/download", params=JAN, headers=admin_headers)
    assert baixado.content == por_bearer.content  # mesmo recorte, mesmo arquivo

    registros = auditoria.listar(2)
    assert [r["status"] for r in registros] == ["ok", "ok"]
    assert {r["usuario"] for r in registros} == {"admin"}
    assert {r["linhas"] for r in registros} == {2}


def test_ticket_de_outro_recorte_nao_baixa_este(cat, client, admin_headers):
    """A amarra do `rec`: com o ticket na URL, editar o filtro na barra de
    endereço não vale — e a recusa vem ANTES de a auditoria abrir."""
    semear_entrada(cat, gem="0000000001")
    ticket = client.post(
        f"{BASE}/download/ticket", params={**JAN, "unidade": ["RMSPII"]}, headers=admin_headers
    ).json()["ticket"]
    antes = len(auditoria.listar(1000))

    r = client.get(f"{BASE}/download", params={**JAN, "ticket": ticket})
    assert r.status_code == 401
    assert "não corresponde ao recorte" in r.json()["detail"]
    assert len(auditoria.listar(1000)) == antes  # download que não saiu não é linha na trilha
