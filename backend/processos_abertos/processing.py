"""Cálculo do resumo semanal de Processos Abertos.

Porta pra Python, com a MESMA lógica de negócio, a função `processFiles()` de
`frontend/processos-abertos/app.js` (linhas 53-81) — o de-para de unidade e de
categoria, o filtro de "em aberto" e o cálculo de dias em aberto continuam
idênticos. Ver docs/PROCESSOS_ABERTOS_MAPA_DADOS_DW.md pra contexto completo
de onde cada relatório entra e o que cada coluna significa.

Particularidades preservadas de propósito (não são bugs a corrigir aqui —
mudar o resultado silenciosamente confundiria quem compara com o histórico já
salvo pelo upload manual):
  - LPN Armazém conta toda linha direto como "> 5 dias", sem olhar data.
  - Recebimento/Expedição recebem contribuição do JDA E do SLIN, somadas.
  - Uma linha com `dias == 0` (evento na própria data de referência) soma no
    total mas não cai em nenhum bucket d1..d5p.
  - O loop do JDA pula unidade vazia; os loops do SLIN não têm essa mesma
    proteção (replicado de propósito, não é uma omissão nova).
"""
from __future__ import annotations

import io
import re
from datetime import date, datetime

from openpyxl import load_workbook
from openpyxl.utils.datetime import from_excel

UNIT_ALIASES: dict[str, str] = {
    "RIO DE JANEIRO": "RMRJ",
    "RMSPII - BARUERI": "RMSPII",
    "SF RPII - RIBEIRAO": "RPII",
    "MAIRINQUE": "MAQ",
    "FORTALEZA": "FOR",
    "RECIFE": "REC",
    "RMSPIV - SANCA": "RMSPV",
    "UNIDADE CURITIBA": "CWBIII",
    "TAC BSB": "BSB",
    "TAMBORÉ": "RMSPII",
}

OP_TO_SG: dict[str, str] = {
    "Descarga": "03 - Portaria em aberto - Recebimento",
    "Cross Docking": "04 - Portaria em aberto - Cross Docking",
    "Distribuição": "04 - Portaria em aberto - Expedição",
}

_RE_DATA = re.compile(r"(\d{1,2})[/-](\d{1,2})[/-](\d{2,4})")
_CAMPOS_BUCKET = ("total", "d1", "d2", "d3", "d4", "d5", "d5p")


class ProcessamentoError(ValueError):
    """Arquivo sem a estrutura esperada (aba/coluna faltando)."""


def canonical_unit(nome: str | None) -> str | None:
    if not nome:
        return None
    return UNIT_ALIASES.get(nome, nome)


def _bucket() -> dict[str, float]:
    return {campo: 0 for campo in _CAMPOS_BUCKET}


def _to_n(valor: object) -> float:
    if valor is None or valor == "":
        return 0.0
    try:
        return float(valor)
    except (TypeError, ValueError):
        return 0.0


def _parse_data(valor: object) -> datetime | None:
    if valor is None or valor == "":
        return None
    if isinstance(valor, datetime):
        return valor
    if isinstance(valor, date):
        return datetime(valor.year, valor.month, valor.day)
    if isinstance(valor, (int, float)):
        try:
            return from_excel(valor)
        except (TypeError, ValueError):
            return None
    m = _RE_DATA.search(str(valor))
    if not m:
        return None
    dia, mes, ano = int(m.group(1)), int(m.group(2)), int(m.group(3))
    if ano < 100:
        ano += 2000
    try:
        return datetime(ano, mes, dia)
    except ValueError:
        return None


def _dias_em_aberto(evento: datetime | None, referencia: datetime) -> int:
    if evento is None:
        return 0
    return max(0, (referencia - evento).days)


def _incrementar_dia(bucket: dict[str, float], dias: int, incremento: float) -> None:
    bucket["total"] += incremento
    if dias == 1:
        bucket["d1"] += incremento
    elif dias == 2:
        bucket["d2"] += incremento
    elif dias == 3:
        bucket["d3"] += incremento
    elif dias == 4:
        bucket["d4"] += incremento
    elif dias == 5:
        bucket["d5"] += incremento
    elif dias > 5:
        bucket["d5p"] += incremento


def _linhas(planilha) -> list[dict]:
    """Equivalente a `XLSX.utils.sheet_to_json`: 1ª linha = cabeçalho."""
    it = planilha.iter_rows(values_only=True)
    try:
        cabecalho_bruto = next(it)
    except StopIteration:
        return []
    cabecalho = [str(c).strip() if c is not None else None for c in cabecalho_bruto]
    linhas: list[dict] = []
    for valores in it:
        if all(v is None for v in valores):
            continue
        linhas.append({c: v for c, v in zip(cabecalho, valores) if c is not None})
    return linhas


def _abrir(conteudo: bytes):
    return load_workbook(io.BytesIO(conteudo), data_only=True, read_only=True)


def montar_semana(slin_bytes: bytes, jda_bytes: bytes) -> dict:
    """Recebe os dois relatórios como bytes e devolve o mesmo formato de
    "semana" que o upload manual monta e envia pro backend hoje."""
    try:
        slin_wb = _abrir(slin_bytes)
    except Exception as exc:
        raise ProcessamentoError(f"não consegui abrir o arquivo do SLIN: {exc}") from exc
    try:
        jda_wb = _abrir(jda_bytes)
    except Exception as exc:
        raise ProcessamentoError(f"não consegui abrir o arquivo do JDA: {exc}") from exc

    if len(slin_wb.worksheets) < 4:
        raise ProcessamentoError(
            f"o arquivo do SLIN devia ter 4 abas (Portaria/Recebimento/Expedição/LPN), "
            f"tem {len(slin_wb.worksheets)}"
        )
    if "Sheet0" not in jda_wb.sheetnames:
        raise ProcessamentoError("o arquivo do JDA não tem a aba 'Sheet0'")

    # Recebimento do SLIN (2ª aba) — acha a data de referência (evento mais recente)
    df2_all = _linhas(slin_wb.worksheets[1])
    ref_date = datetime(2000, 1, 1)
    for linha in df2_all:
        evento = _parse_data(linha.get("dthr_chegada"))
        if evento and evento > ref_date:
            ref_date = evento
    ref_str = ref_date.strftime("%d/%m/%Y")

    jda_s0 = _linhas(jda_wb["Sheet0"])
    df1_all = _linhas(slin_wb.worksheets[0])
    port_open = [r for r in df1_all if r.get("status_portaria") != "Concluído"]
    df2s = df2_all
    df3s = _linhas(slin_wb.worksheets[2])
    df4s = _linhas(slin_wb.worksheets[3])

    resumo: dict[str, dict] = {}
    tipos: dict[str, dict] = {}

    # JDA: já vem bucketizado por dias — só soma.
    for row in jda_s0:
        u = canonical_unit(row.get("Unidade"))
        if not u:
            continue
        r = resumo.setdefault(u, _bucket())
        r["total"] += _to_n(row.get("Qtd Total"))
        r["d1"] += _to_n(row.get("1D"))
        r["d2"] += _to_n(row.get("2D"))
        r["d3"] += _to_n(row.get("3D"))
        r["d4"] += _to_n(row.get("4D"))
        r["d5"] += _to_n(row.get("5D"))
        r["d5p"] += _to_n(row.get("> 5D"))

        sg = row.get("Sub Grupo")
        if not sg:
            continue
        t = tipos.setdefault(sg, _bucket())
        t["total"] += _to_n(row.get("Qtd Total"))
        t["d1"] += _to_n(row.get("1D"))
        t["d2"] += _to_n(row.get("2D"))
        t["d3"] += _to_n(row.get("3D"))
        t["d4"] += _to_n(row.get("4D"))
        t["d5"] += _to_n(row.get("5D"))
        t["d5p"] += _to_n(row.get("> 5D"))

    # Portaria aberta do SLIN: de-para de unidade + cálculo de dias na mão.
    for row in port_open:
        u = canonical_unit(row.get("nome_und"))
        r = resumo.setdefault(u, _bucket())
        dias = _dias_em_aberto(_parse_data(row.get("dthr_chegada")), ref_date)
        _incrementar_dia(r, dias, 1)

        sg = OP_TO_SG.get(row.get("operacao"))
        if not sg:
            continue
        t = tipos.setdefault(sg, _bucket())
        _incrementar_dia(t, dias, 1)

    # Recebimento + Expedição do SLIN: mesma conta, categoria fixa por aba
    # (soma com o que o JDA já trouxe pra essas mesmas categorias, de propósito).
    for linhas_aba, sg in ((df2s, "01 - Recebimento em aberto"), (df3s, "02 - Expedição em aberto")):
        for row in linhas_aba:
            u = canonical_unit(row.get("nome_und"))
            r = resumo.setdefault(u, _bucket())
            dias = _dias_em_aberto(_parse_data(row.get("dthr_chegada")), ref_date)
            _incrementar_dia(r, dias, 1)

            t = tipos.setdefault(sg, _bucket())
            _incrementar_dia(t, dias, 1)

    # LPN Armazém: bloco só, sempre "> 5 dias" — não tem data pra calcular.
    if df4s:
        t = tipos.setdefault("06 - LPN Armazém", _bucket())
        t["total"] += len(df4s)
        t["d5p"] += len(df4s)
        for row in df4s:
            u = canonical_unit(row.get("nome_und"))
            r = resumo.setdefault(u, _bucket())
            r["total"] += 1
            r["d5p"] += 1

    gtotal = sum(v["total"] for v in resumo.values())
    gd5p = sum(v["d5p"] for v in resumo.values())
    gd1 = sum(v["d1"] for v in resumo.values())
    gd25 = sum(v["d2"] + v["d3"] + v["d4"] + v["d5"] for v in resumo.values())

    return {
        "date": ref_str,
        "total": int(round(gtotal)),
        "d5p": int(round(gd5p)),
        "d1": int(round(gd1)),
        "d25": int(round(gd25)),
        # Original (app.js) manda `null` quando gtotal é 0 — inatingível na prática
        # e o backend não aceita `pct` nulo; 0.0 é o fallback seguro aqui.
        "pct": round(gd5p / gtotal * 100, 1) if gtotal else 0.0,
        "units": len(resumo),
        "resumo": resumo,
        "tipos": tipos,
    }
